import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def crear_formulario():
    # 1. Crear el libro de trabajo y configurar la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitud de Reducción"
    
    # Asegurar que se vean las líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # 2. Definición de Estilos Visuales (Paleta Corporativa)
    font_titulo = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    font_seccion = Font(name="Arial", size=12, bold=True, color="1B365D")
    font_subseccion = Font(name="Arial", size=10, bold=True, color="555555")
    font_header_tabla = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_datos = Font(name="Arial", size=10, bold=False)
    font_calculos = Font(name="Arial", size=10, bold=True, color="1B365D")
    font_alerta = Font(name="Arial", size=10, bold=True, color="9C0006")

    # Fills (Relleno de Celdas)
    fill_azul_oscuro = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_azul_claro = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    fill_gris_tabla = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    fill_editable = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid") # Amarillo pastel
    fill_calculo_tabla = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Verde menta pastel

    # Bordes
    borde_fino = Side(border_style="thin", color="D1D5DB")
    border_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    
    borde_editable = Side(border_style="medium", color="FBC02D")
    border_input = Border(left=borde_editable, right=borde_editable, top=borde_editable, bottom=borde_editable)

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # 3. TÍTULO PRINCIPAL
    ws.merge_cells("A1:G2")
    ws["A1"] = "SOLICITUD DE REDUCCIÓN DE JORNADA LABORAL"
    ws["A1"].font = font_titulo
    ws["A1"].fill = fill_azul_oscuro
    ws["A1"].alignment = align_center

    # 4. DATOS GENERALES DEL EMPLEADO (Campos Editables)
    ws["A4"] = "DATOS GENERALES"
    ws["A4"].font = font_seccion
    
    datos_personales = [
        ("B5", "Nombre del Empleado:", "C5"),
        ("B6", "Departamento / Área:", "C6"),
        ("B7", "Fecha de Solicitud:", "C7"),
    ]
    for lbl_celda, texto, val_celda in datos_personales:
        ws[lbl_celda] = texto
        ws[lbl_celda].font = font_subseccion
        ws[lbl_celda].alignment = align_right
        
        ws[val_celda].fill = fill_editable
        ws[val_celda].border = border_input

    # 5. CONFIGURACIÓN DEL PERIODO SOLICITADO (Campos Editables / Selectores)
    ws["E4"] = "PERIODOS DISPONIBLES"
    ws["E4"].font = font_seccion
    
    periodos = [
        ("E5", "Solicitar Invierno (8 meses):", "F5", "SI"),
        ("E6", "Solicitar Verano (4 meses):", "F6", "SI"),
    ]
    for lbl_celda, texto, val_celda, por_defecto in periodos:
        ws[lbl_celda] = texto
        ws[lbl_celda].font = font_subseccion
        ws[lbl_celda].alignment = align_right
        
        ws[val_celda] = por_defecto
        ws[val_celda].fill = fill_editable
        ws[val_celda].border = border_input
        ws[val_celda].alignment = align_center

    # 6. TABLA PRINCIPAL DE PLANIFICACIÓN HORARIA Y REDUCCIÓN
    ws["A9"] = "PLANIFICACIÓN HORARIA POR PERIODO (INVIERNO / VERANO)"
    ws["A9"].font = font_seccion

    headers = [
        "Parámetro / Concepto", "Invierno (8 Meses)", "Verano (4 Meses)", 
        "Cálculo Ponderado Anual", "Instrucciones de Uso"
    ]
    
    ws.merge_cells("A11:B11")
    ws["A11"] = headers[0]
    ws["C11"] = headers[1]
    ws["D11"] = headers[2]
    ws["E11"] = headers[3]
    ws.merge_cells("F11:G11")
    ws["F11"] = headers[4]

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        celda = ws[f"{col}11"]
        celda.font = font_header_tabla
        celda.fill = fill_gris_tabla
        celda.alignment = align_center

    # Estructura de Filas de la Tabla Principal
    # Formato: (Fila, Concepto, Es_Editable_Invi, Es_Editable_Ver, Valor_Def_Invi, Valor_Def_Ver, Formula_Pond, Formato_Num)
    tabla_estructura = [
        (12, "Modo de Entrada (Horario / Manual)", True, True, "Horario", "Horario", "", "@"),
        (13, "NRA - Horas Base Semanales del Puesto", True, True, 40.0, 36.0, "=IF(AND(F5=\"SI\",F6=\"SI\"),(C13*8/12)+(D13*4/12),IF(F5=\"SI\",C13,IF(F6=\"SI\",D13,0)))", "0.00"),
        (14, "Hora Entrada (Modo Horario)", True, True, "08:00", "08:00", "", "hh:mm"),
        (15, "Hora Salida (Modo Horario)", True, True, "17:00", "15:00", "", "hh:mm"),
        (16, "Descuento Comida Automático (Horas)", True, True, 1.0, 0.0, "", "0.00"),
        (17, "Horas Diarias Introducidas (Modo Manual)", True, True, 0.0, 0.0, "", "0.00"),
        (18, "Total Horas Diarias Calculadas", False, False, "=IF(C12=\"Horario\",(C15-C14)*24-C16,C17)", "=IF(D12=\"Horario\",(D15-D14)*24-D16,D17)", "", "0.00"),
        (19, "Total Horas Semanales Propuestas", False, False, "=C18*5", "=D18*5", "=IF(AND(F5=\"SI\",F6=\"SI\"),(C19*8/12)+(D19*4/12),IF(F5=\"SI\",C19,IF(F6=\"SI\",D19,0)))", "0.00"),
        (20, "% Jornada Resultante", False, False, "=(C19/C13)", "=(D19/D13)", "=IF(AND(F5=\"SI\",F6=\"SI\"),(C20*8/12)+(D20*4/12),IF(F5=\"SI\",C20,IF(F6=\"SI\",D20,0)))", "0.0%"),
        (21, "% Reducción de Jornada", False, False, "=1-C20", "=1-D20", "=IF(AND(F5=\"SI\",F6=\"SI\"),(C21*8/12)+(D21*4/12),IF(F5=\"SI\",C21,IF(F6=\"SI\",D21,0)))", "0.0%"),
        (22, "Coeficiente de Jornada Reducida", False, False, "=C20", "=D20", "=IF(AND(F5=\"SI\",F6=\"SI\"),(C22*8/12)+(D22*4/12),IF(F5=\"SI\",C22,IF(F6=\"SI\",D22,0)))", "0.0000"),
        (23, "Salario Bruto Mensual Ordinario (Base)", True, True, 2500.0, 2500.0, "=IF(AND(F5=\"SI\",F6=\"SI\"),(C23*8/12)+(D23*4/12),IF(F5=\"SI\",C23,IF(F6=\"SI\",D23,0)))", "#,##0.00 €"),
        (24, "Retribución Bruta Estimada (RBA Ajustada)", False, False, "=C23*C22", "=D23*D22", "=(C24*8)+(D24*4)", "#,##0.00 €")
    ]

    instrucciones = {
        12: "Escribe 'Horario' o 'Manual'",
        13: "Horas de contrato completo del puesto",
        14: "Formato 24h (Ej. 08:30)",
        15: "Formato 24h (Ej. 17:30)",
        16: "Tiempo de comida a restar en horas",
        17: "Solo si usas modo 'Manual'",
        18: "Cálculo según el modo elegido",
        19: "Horas diarias multiplicadas por 5 días",
        20: "Proporción de jornada trabajada",
        21: "Porcentaje reducido del total original",
        22: "Multiplicador aplicable al salario",
        23: "Salario mensual bruto al 100% de jornada",
        24: "Salario mensual ajustado / Col E es Total Anual"
    }

    # Rellenar filas y aplicar estilos
    for fila, concepto, edit_invi, edit_ver, def_invi, def_ver, form_pond, num_fmt in tabla_estructura:
        ws.merge_cells(f"A{fila}:B{fila}")
        ws[f"A{fila}"] = concepto
        ws[f"A{fila}"].font = font_datos
        ws[f"A{fila}"].alignment = align_left
        
        # Columna Invierno
        ws[f"C{fila}"] = def_invi
        ws[f"C{fila}"].number_format = num_fmt
        if edit_invi:
            ws[f"C{fila}"].fill = fill_editable
            ws[f"C{fila}"].border = border_input
            ws[f"C{fila}"].font = font_datos
            ws[f"C{fila}"].alignment = align_center if isinstance(def_invi, str) else align_right
        else:
            ws[f"C{fila}"].fill = fill_calculo_tabla
            ws[f"C{fila}"].border = border_celda
            ws[f"C{fila}"].font = font_calculos
            ws[f"C{fila}"].alignment = align_right

        # Columna Verano
        ws[f"D{fila}"] = def_ver
        ws[f"D{fila}"].number_format = num_fmt
        if edit_ver:
            ws[f"D{fila}"].fill = fill_editable
            ws[f"D{fila}"].border = border_input
            ws[f"D{fila}"].font = font_datos
            ws[f"D{fila}"].alignment = align_center if isinstance(def_ver, str) else align_right
        else:
            ws[f"D{fila}"].fill = fill_calculo_tabla
            ws[f"D{fila}"].border = border_celda
            ws[f"D{fila}"].font = font_calculos
            ws[f"D{fila}"].alignment = align_right

        # Columna Ponderada Anual
        if form_pond:
            ws[f"E{fila}"] = form_pond
            ws[f"E{fila}"].number_format = num_fmt
            ws[f"E{fila}"].fill = fill_azul_claro
            ws[f"E{fila}"].font = font_calculos
            ws[f"E{fila}"].border = border_celda
            ws[f"E{fila}"].alignment = align_right
        else:
            ws[f"E{fila}"] = "-"
            ws[f"E{fila}"].fill = fill_azul_claro
            ws[f"E{fila}"].alignment = align_center
            ws[f"E{fila}"].border = border_celda

        # Columna Instrucciones
        ws.merge_cells(f"F{fila}:G{fila}")
        ws[f"F{fila}"] = instrucciones.get(fila, "")
        ws[f"F{fila}"].font = font_subseccion
        ws[f"F{fila}"].alignment = align_left
        
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            if not ws[f"{col}{fila}"].border:
                ws[f"{col}{fila}"].border = border_celda

    # 7. CUADRO DE VALIDACIÓN LEGAL (12,5% - 50%)
    ws["A26"] = "VALIDACIÓN LEGAL DE LA SOLICITUD"
    ws["A26"].font = font_seccion

    ws.merge_cells("A27:C27")
    ws["A27"] = "Estado Validación Reducción Invierno (12.5% - 50%):"
    ws["A27"].font = font_datos
    ws["A27"].alignment = align_right
    ws["D27"] = '=IF(F5="NO","NO SOLICITADO",IF(AND(C21>=0.125,C21<=0.5),"CORRECTO","ERROR: Fuera de rango legal"))'
    ws["D27"].font = font_alerta
    ws["D27"].fill = fill_calculo_tabla
    ws["D27"].border = border_celda
    ws["D27"].alignment = align_center

